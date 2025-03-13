from openpyxl import Workbook, load_workbook


# 写入新文件
def write_new_excel(filename, sheet_name, cell, text):
    # 创建新工作簿
    wb = Workbook()
    # 选择活动工作表或新建指定名称的工作表
    ws = wb.active
    ws.title = sheet_name  # 重命名工作表
    # 写入单元格
    ws[cell] = text
    # 保存文件
    wb.save(filename)
    print(f"成功写入新文件 {filename}")


# 修改现有文件
def update_existing_excel(filename, sheet_name, cell, text):
    # 加载现有工作簿
    wb = load_workbook(filename)
    # 检查工作表是否存在，不存在则创建
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    # 写入单元格
    ws[cell] = text
    # 保存文件（覆盖原文件）
    wb.save(filename)
    print(f"成功更新文件 {filename}")

# 写入新文件
def write_excel(filename, sheet_name, cell, text,wb):

    # 选择活动工作表或新建指定名称的工作表
    ws = wb.active
    ws.title = sheet_name  # 重命名工作表
    # 写入单元格
    ws[cell] = text
    # 保存文件
    wb.save(filename)
    print(f"成功写入新文件 {filename}")
# 使用示例
if __name__ == "__main__":
    # 写入新文件
    write_new_excel("test.xlsx", "Sheet1", "A1", "Hello World")

    # 修改现有文件（假设文件已存在）
    update_existing_excel("test.xlsx", "Sheet1", "A2", "Python写入示例")